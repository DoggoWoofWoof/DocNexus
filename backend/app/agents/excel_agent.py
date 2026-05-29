import json
import re
import sys
from collections.abc import Callable
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlmodel import Session

from backend.app.core.config import Settings
from backend.app.schemas.artifact import ArtifactRef, ArtifactType
from backend.app.schemas.physician import PhysicianRead
from backend.app.services.artifacts import finalize_artifact_file, register_artifact, to_artifact_ref
from backend.app.services.artifact_workers import run_artifact_worker
from backend.app.services.prompts import load_prompt


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill(fill_type="solid", fgColor="EAF2F8")
TextGenerator = Callable[[list[dict[str, object]]], str]


def generate_excel_workbook(
    *,
    session: Session,
    settings: Settings,
    generate_text: TextGenerator,
    analysis_type: str,
    physicians: list[PhysicianRead],
    dimensions: list[str],
    icd10_codes: list[str],
    artifact_provenance: dict[str, object] | None = None,
) -> ArtifactRef:
    plan = _build_workbook_plan(
        generate_text=generate_text,
        analysis_type=analysis_type,
        physicians=physicians,
        dimensions=dimensions,
        icd10_codes=icd10_codes,
    )
    filename = f"{_slugify(analysis_type or 'physician_breakdown')}.xlsx"
    artifact, path = register_artifact(
        session,
        settings=settings,
        artifact_type=ArtifactType.xlsx,
        filename=filename,
        source_agent="excel",
        **_with_plan(artifact_provenance, plan),
    )

    render_metadata = run_artifact_worker(
        settings=settings,
        module="backend.app.agents.excel_agent",
        payload={
            "plan": plan,
            "physicians": [physician.model_dump(by_alias=True) for physician in physicians],
            "icd10Codes": icd10_codes,
        },
        output_path=path,
        e2b_script=_e2b_render_script(),
        packages=["openpyxl"],
    )
    _store_render_metadata(session, artifact, render_metadata)
    finalize_artifact_file(session, artifact, path)
    return to_artifact_ref(artifact)


def _render_workbook(
    *,
    path: Path,
    plan: dict[str, Any],
    physicians: list[PhysicianRead],
    icd10_codes: list[str],
) -> None:
    workbook = Workbook()
    workbook.properties.title = str(plan["title"])
    workbook.properties.subject = str(plan["summary"])
    workbook.properties.keywords = ", ".join(_string_list(plan["analysisNotes"])[:5])
    raw_sheet = workbook.active
    raw_sheet.title = "Raw Physician Data"
    _write_raw_data(raw_sheet, physicians, icd10_codes)

    summary_sheet = workbook.create_sheet("State x Specialty Summary")
    _write_state_specialty_summary(summary_sheet, physicians)

    icd_sheet = workbook.create_sheet("ICD-10 Breakdown")
    _write_icd10_breakdown(icd_sheet, physicians, icd10_codes)

    for sheet in workbook.worksheets:
        _format_sheet(sheet)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _write_raw_data(sheet: Worksheet, physicians: list[PhysicianRead], icd10_codes: list[str]) -> None:
    codes = _selected_or_all_codes(physicians, icd10_codes)
    headers = [
        "ID",
        "NPI",
        "First Name",
        "Last Name",
        "Specialty",
        "Affiliation",
        "City",
        "State",
        "Total NSCLC Claims",
        "Volume Tier",
        "Email",
        "Board Certified",
        *codes,
    ]
    sheet.append(headers)

    for physician in physicians:
        sheet.append(
            [
                physician.id,
                physician.npi,
                physician.firstName,
                physician.lastName,
                physician.specialty,
                physician.affiliation,
                physician.city,
                physician.state,
                physician.totalNSCLCClaims,
                physician.volumeTier,
                physician.email,
                "Yes" if physician.boardCertified else "No",
                *[physician.icd10ClaimVolume.get(code, 0) for code in codes],
            ]
        )


def _write_state_specialty_summary(sheet: Worksheet, physicians: list[PhysicianRead]) -> None:
    sheet.append(["State", "Specialty", "Physician Count", "Total NSCLC Claims", "Average NSCLC Claims"])

    grouped: dict[tuple[str, str], list[PhysicianRead]] = defaultdict(list)
    for physician in physicians:
        grouped[(physician.state, physician.specialty)].append(physician)

    rows = []
    for (state, specialty), group in grouped.items():
        total_claims = sum(physician.totalNSCLCClaims for physician in group)
        rows.append(
            [
                state,
                specialty,
                len(group),
                total_claims,
                round(total_claims / len(group), 1),
            ]
        )

    for row in sorted(rows, key=lambda item: (item[0], item[1])):
        sheet.append(row)


def _write_icd10_breakdown(sheet: Worksheet, physicians: list[PhysicianRead], icd10_codes: list[str]) -> None:
    codes = _selected_or_all_codes(physicians, icd10_codes)
    sheet.append(["ICD-10 Code", "Physician Count", "Total Claims", "Average Claims Per Physician"])

    for code in codes:
        claim_values = [physician.icd10ClaimVolume.get(code, 0) for physician in physicians]
        non_zero = [value for value in claim_values if value > 0]
        total = sum(claim_values)
        average = round(total / len(non_zero), 1) if non_zero else 0
        sheet.append([code, len(non_zero), total, average])


def _selected_or_all_codes(physicians: list[PhysicianRead], icd10_codes: list[str]) -> list[str]:
    if icd10_codes:
        return sorted({code.upper() for code in icd10_codes})

    codes: set[str] = set()
    for physician in physicians:
        codes.update(physician.icd10ClaimVolume.keys())
    return sorted(codes)


def _format_sheet(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if row_number % 2 == 0:
            for cell in row:
                cell.fill = ALT_FILL

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)


def _slugify(value: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", value.strip().lower()).strip("_")
    return slug[:80] or "physician_breakdown"


def _build_workbook_plan(
    *,
    generate_text: TextGenerator,
    analysis_type: str,
    physicians: list[PhysicianRead],
    dimensions: list[str],
    icd10_codes: list[str],
) -> dict[str, Any]:
    fallback = _fallback_workbook_plan(analysis_type, physicians, dimensions, icd10_codes)
    messages: list[dict[str, object]] = [
        {"role": "system", "content": load_prompt("excel_agent.md")},
        {
            "role": "user",
            "content": json.dumps(
                {
                    "analysisType": analysis_type,
                    "dimensions": dimensions,
                    "icd10Codes": icd10_codes,
                    "physicianCount": len(physicians),
                    "physicians": [physician.model_dump(by_alias=True) for physician in physicians],
                    "requiredOutput": {
                        "title": "Workbook title",
                        "summary": "One-sentence workbook purpose",
                        "sheetPlan": "List of sheet names and purpose",
                        "analysisNotes": "3-5 grounded notes for reviewer context",
                    },
                }
            ),
        },
    ]
    try:
        raw_plan = json.loads(_strip_code_fences(generate_text(messages)))
    except Exception:
        return fallback
    if not isinstance(raw_plan, dict):
        return fallback

    sheet_plan = raw_plan.get("sheetPlan")
    if not isinstance(sheet_plan, list):
        sheet_plan = fallback["sheetPlan"]
    analysis_notes = _string_list(raw_plan.get("analysisNotes"))
    if len(analysis_notes) < 2:
        analysis_notes = fallback["analysisNotes"]

    return {
        "title": str(raw_plan.get("title") or fallback["title"])[:120],
        "summary": str(raw_plan.get("summary") or fallback["summary"])[:240],
        "sheetPlan": sheet_plan,
        "analysisNotes": analysis_notes[:5],
    }


def _fallback_workbook_plan(
    analysis_type: str,
    physicians: list[PhysicianRead],
    dimensions: list[str],
    icd10_codes: list[str],
) -> dict[str, Any]:
    total_claims = sum(physician.totalNSCLCClaims for physician in physicians)
    return {
        "title": analysis_type or "Physician Breakdown",
        "summary": f"Workbook covering {len(physicians)} physicians and {total_claims} total NSCLC claims.",
        "sheetPlan": [
            {"name": "Raw Physician Data", "purpose": "Filtered physician-level rows."},
            {"name": "State x Specialty Summary", "purpose": "Pivot-style claim volume summary."},
            {"name": "ICD-10 Breakdown", "purpose": "Claim volume by selected ICD-10 code."},
        ],
        "analysisNotes": [
            f"Dimensions requested: {', '.join(dimensions) if dimensions else 'state, specialty, ICD-10 code'}.",
            f"ICD-10 scope: {', '.join(icd10_codes) if icd10_codes else 'all available NSCLC codes'}.",
            "Raw data is preserved before aggregation.",
        ],
    }


def _with_plan(
    artifact_provenance: dict[str, object] | None,
    plan: dict[str, Any],
) -> dict[str, object]:
    updated = dict(artifact_provenance or {})
    provenance = dict(updated.get("provenance") or {})
    provenance["llmPlan"] = plan
    updated["provenance"] = provenance
    return updated


def _store_render_metadata(session: Session, artifact, render_metadata: dict[str, object]) -> None:
    provenance = dict(artifact.provenance or {})
    provenance["renderExecution"] = render_metadata
    artifact.provenance = provenance
    session.add(artifact)
    session.commit()
    session.refresh(artifact)


def _e2b_render_script() -> str:
    return r'''
import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill(fill_type="solid", fgColor="EAF2F8")


def main() -> int:
    payload = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
    output_path = Path(sys.argv[2])
    plan = payload["plan"]
    physicians = payload["physicians"]
    icd10_codes = payload.get("icd10Codes") or []

    workbook = Workbook()
    workbook.properties.title = str(plan.get("title") or "Physician Breakdown")
    workbook.properties.subject = str(plan.get("summary") or "")
    workbook.properties.keywords = ", ".join(str(note) for note in (plan.get("analysisNotes") or [])[:5])

    raw_sheet = workbook.active
    raw_sheet.title = "Raw Physician Data"
    write_raw_data(raw_sheet, physicians, icd10_codes)

    summary_sheet = workbook.create_sheet("State x Specialty Summary")
    write_state_specialty_summary(summary_sheet, physicians)

    icd_sheet = workbook.create_sheet("ICD-10 Breakdown")
    write_icd10_breakdown(icd_sheet, physicians, icd10_codes)

    for sheet in workbook.worksheets:
        format_sheet(sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return 0


def write_raw_data(sheet, physicians, icd10_codes):
    codes = selected_or_all_codes(physicians, icd10_codes)
    headers = [
        "ID",
        "NPI",
        "First Name",
        "Last Name",
        "Specialty",
        "Affiliation",
        "City",
        "State",
        "Total NSCLC Claims",
        "Volume Tier",
        "Email",
        "Board Certified",
        *codes,
    ]
    sheet.append(headers)
    for physician in physicians:
        volumes = physician.get("icd10ClaimVolume") or {}
        sheet.append(
            [
                physician.get("id"),
                physician.get("npi"),
                physician.get("firstName"),
                physician.get("lastName"),
                physician.get("specialty"),
                physician.get("affiliation"),
                physician.get("city"),
                physician.get("state"),
                physician.get("totalNSCLCClaims"),
                physician.get("volumeTier"),
                physician.get("email"),
                "Yes" if physician.get("boardCertified") else "No",
                *[volumes.get(code, 0) for code in codes],
            ]
        )


def write_state_specialty_summary(sheet, physicians):
    sheet.append(["State", "Specialty", "Physician Count", "Total NSCLC Claims", "Average NSCLC Claims"])
    grouped = defaultdict(list)
    for physician in physicians:
        grouped[(physician.get("state"), physician.get("specialty"))].append(physician)
    rows = []
    for (state, specialty), group in grouped.items():
        total_claims = sum(int(physician.get("totalNSCLCClaims") or 0) for physician in group)
        rows.append([state, specialty, len(group), total_claims, round(total_claims / len(group), 1)])
    for row in sorted(rows, key=lambda item: (str(item[0]), str(item[1]))):
        sheet.append(row)


def write_icd10_breakdown(sheet, physicians, icd10_codes):
    codes = selected_or_all_codes(physicians, icd10_codes)
    sheet.append(["ICD-10 Code", "Physician Count", "Total Claims", "Average Claims Per Physician"])
    for code in codes:
        claim_values = [int((physician.get("icd10ClaimVolume") or {}).get(code, 0)) for physician in physicians]
        non_zero = [value for value in claim_values if value > 0]
        total = sum(claim_values)
        average = round(total / len(non_zero), 1) if non_zero else 0
        sheet.append([code, len(non_zero), total, average])


def selected_or_all_codes(physicians, icd10_codes):
    if icd10_codes:
        return sorted({str(code).upper() for code in icd10_codes})
    codes = set()
    for physician in physicians:
        codes.update((physician.get("icd10ClaimVolume") or {}).keys())
    return sorted(codes)


def format_sheet(sheet):
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if row_number % 2 == 0:
            for cell in row:
                cell.fill = ALT_FILL
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)


if __name__ == "__main__":
    raise SystemExit(main())
'''.strip()


def _string_list(value: object) -> list[str]:
    if not isinstance(value, list):
        return []
    return [str(item).strip() for item in value if str(item).strip()]


def _strip_code_fences(text: str) -> str:
    stripped = text.strip()
    if stripped.startswith("```"):
        lines = stripped.splitlines()
        if lines and lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].startswith("```"):
            lines = lines[:-1]
        return "\n".join(lines)
    return stripped


def _render_worker(argv: list[str]) -> int:
    if len(argv) != 4 or argv[1] != "--render-worker":
        return 2

    payload_path = Path(argv[2])
    output_path = Path(argv[3])
    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    physicians = [PhysicianRead.model_validate(item) for item in payload["physicians"]]
    _render_workbook(
        path=output_path,
        plan=payload["plan"],
        physicians=physicians,
        icd10_codes=list(payload.get("icd10Codes") or []),
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(_render_worker(sys.argv))
